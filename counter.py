from collections import Counter
import datetime
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import fitz
import os


def process_pdf_file(name, keywords):
    doc = fitz.open("assets/{}.pdf".format(name))
    plainText = ''
    for page in doc:
        plainText += page.get_text()
    word_list = plainText.split()
    char_counts = Counter(word_list)
    keyword_counts = {keyword: char_counts[keyword] for keyword in keywords}
    return (name, keyword_counts)


def process_pdf_files_in_parallel(filenames, keywords, max_workers):
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = []
        for name in filenames:
            future = executor.submit(process_pdf_file, name, keywords)
            futures.append(future)

        results = []
        for future in futures:
            result = future.result()
            results.append(result)

        return results


def generate_heatmap_pdf_file(name):
    df = pd.read_excel('outputs/keywords/'+name)
    df = df.drop(columns=['Technology areas'])
    df = df.set_index('Technologies')
    df = df.transpose()

    plt.figure(figsize=(15, 10))

    # Generate the heatmap using Seaborn
    # colormap = sns.color_palette("Blues")
    sns.heatmap(df, cmap='YlOrRd')

    # # Add labels and title
    plt.title('Keyword counts per publisher')
    plt.ylabel('Publishers or Filenames')
    plt.xlabel('Keywords')
    plt.tight_layout()

    # # Show the plot
    # plt.show()

    # Save the figure to a PDF file
    now = datetime.datetime.now()
    name = 'heatmap_'+now.strftime("%y%m%d%H%M%S")+'.pdf'
    plt.savefig('outputs/heatmaps/'+name)


def main():
    excelfile = load_workbook('Keyword Lists.xlsx')
    sheet = excelfile.active
    path = "assets"
    filenames = []
    for filename in os.listdir(path):
        if os.path.isfile(os.path.join(path, filename)):
            name, extension = os.path.splitext(filename)
            if extension == ".pdf":
                filenames.append(name)

    keywords = []
    for row in range(2, sheet.max_row + 1):
        keywords.append(sheet['B' + str(row)].value)

    pageColRound = ''
    pageColIndex = 'C'
    results = process_pdf_files_in_parallel(filenames, keywords, max_workers=4)
    for i, result in enumerate(results):
        name, keyword_counts = result
        sheet[pageColRound + pageColIndex + '1'] = name
        index = 2
        for v in keyword_counts.values():
            sheet[pageColRound + pageColIndex + str(index)] = v
            index += 1
        if pageColRound == '' and pageColIndex == 'Z':
            pageColRound = 'A'
            pageColIndex = 'A'
        elif pageColRound != '' and pageColIndex == 'Z':
            pageColRound = chr(ord(pageColRound) + 1)
            pageColIndex = 'A'
        else:
            pageColIndex = chr(ord(pageColIndex) + 1)
    now = datetime.datetime.now()
    name = 'keyword_count_' + now.strftime("%y%m%d%H%M%S")+'.xlsx'
    excelfile.save('outputs/keywords/'+name)
    generate_heatmap_pdf_file(name)


if __name__ == '__main__':
    main()
